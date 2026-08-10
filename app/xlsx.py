"""Build a small Excel workbook in memory."""
from io import BytesIO
from typing import Iterable, Sequence

MIN_WIDTH = 12
MAX_WIDTH = 80


def workbook_bytes(
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    sheet_title: str = "Sheet1",
) -> bytes:
    """One sheet, a bold header row, and columns sized to their content."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    # Excel refuses some characters in a sheet name and truncates at 31.
    sheet.title = "".join(c for c in sheet_title if c not in r"[]:*?/\\")[:31] or "Sheet1"

    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    widths = [len(str(header)) for header in headers]

    for row in rows:
        values = ["" if value is None else str(value) for value in row]
        sheet.append(values)
        for index, value in enumerate(values):
            if index < len(widths):
                widths[index] = max(widths[index], len(value))

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(
            MIN_WIDTH, min(width + 2, MAX_WIDTH)
        )

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()
