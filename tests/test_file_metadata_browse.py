"""Browsing and exporting the whole `file_metadata` table."""
import io

from conftest import minimal_patient


def _application(client, **patient_overrides):
    patient_id = client.post(
        "/patients", json=minimal_patient(**patient_overrides)
    ).json()["id"]
    application_id = client.post(
        "/applications", json={"patient_id": patient_id}
    ).json()["id"]
    return patient_id, application_id


def _upload(client, application_id, name="scan.pdf", data=b"%PDF-1.4 fake"):
    return client.post(
        f"/applications/{application_id}/files",
        files=[("files", (name, data, "application/pdf"))],
    ).json()[0]


def test_every_extraction_is_listed_with_its_file(as_admin, storage_root):
    patient_id, application_id = _application(as_admin)
    record = _upload(as_admin, application_id)

    rows = as_admin.get("/file-metadata").json()

    assert len(rows) == 1
    row = rows[0]
    assert row["file_id"] == record["id"]
    assert row["file_name"] == "scan.pdf"
    assert row["application_id"] == application_id
    assert row["patient_id"] == patient_id
    assert isinstance(row["metadata"], dict)


def test_search_reaches_inside_the_stored_metadata(as_admin, storage_root, cursor):
    _, application_id = _application(as_admin)
    first = _upload(as_admin, application_id, name="alpha.pdf")
    _upload(as_admin, application_id, name="beta.pdf")

    # Put a value on one of them that appears nowhere else.
    for row in cursor.store["file_metadata"]:
        if row["file_id"] == first["id"]:
            row["metadata"] = '{"Manufacturer": "Siemens Healthineers"}'

    hits = as_admin.get("/file-metadata", params={"search": "siemens"}).json()

    assert [row["file_id"] for row in hits] == [first["id"]]


def test_search_also_matches_a_field_name(as_admin, storage_root, cursor):
    _, application_id = _application(as_admin)
    record = _upload(as_admin, application_id)

    for row in cursor.store["file_metadata"]:
        if row["file_id"] == record["id"]:
            row["metadata"] = '{"PatientBirthDate": "19700101"}'

    hits = as_admin.get(
        "/file-metadata", params={"search": "patientbirthdate"}
    ).json()

    assert len(hits) == 1


def test_search_matches_the_file_name(as_admin, storage_root):
    _, application_id = _application(as_admin)
    _upload(as_admin, application_id, name="consent-form.pdf")
    _upload(as_admin, application_id, name="referral.pdf")

    hits = as_admin.get("/file-metadata", params={"search": "consent"}).json()

    assert [row["file_name"] for row in hits] == ["consent-form.pdf"]


def test_filters_narrow_by_status_and_type(as_admin, storage_root):
    _, application_id = _application(as_admin)
    _upload(as_admin, application_id)

    stored = as_admin.get("/file-metadata").json()[0]

    kept = as_admin.get("/file-metadata", params={"status": stored["status"]}).json()
    assert len(kept) == 1

    other = "ok" if stored["status"] != "ok" else "unsupported"
    assert as_admin.get("/file-metadata", params={"status": other}).json() == []

    by_type = as_admin.get(
        "/file-metadata", params={"file_type": stored["file_type"].upper()}
    ).json()
    assert len(by_type) == 1
    assert as_admin.get("/file-metadata", params={"file_type": "dicom"}).json() == []


def test_the_export_is_a_real_workbook_of_the_filtered_rows(
    as_admin, storage_root, cursor
):
    from openpyxl import load_workbook

    _, application_id = _application(as_admin)
    first = _upload(as_admin, application_id, name="alpha.pdf")
    _upload(as_admin, application_id, name="beta.pdf")

    for row in cursor.store["file_metadata"]:
        if row["file_id"] == first["id"]:
            row["metadata"] = '{"Manufacturer": "Siemens"}'

    response = as_admin.get(
        "/file-metadata/export", params={"search": "siemens"}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )
    assert "file-metadata-" in response.headers["content-disposition"]

    sheet = load_workbook(io.BytesIO(response.content)).active
    headers = [cell.value for cell in sheet[1]]

    assert headers[0] == "File"
    # The extracted fields become columns of their own.
    assert "Manufacturer" in headers

    body = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(body) == 1
    assert body[0][0] == "alpha.pdf"
    assert body[0][headers.index("Manufacturer")] == "Siemens"


def test_the_export_covers_every_row_when_nothing_is_searched_for(
    as_admin, storage_root
):
    from openpyxl import load_workbook

    _, application_id = _application(as_admin)
    _upload(as_admin, application_id, name="alpha.pdf")
    _upload(as_admin, application_id, name="beta.pdf")

    response = as_admin.get("/file-metadata/export")
    sheet = load_workbook(io.BytesIO(response.content)).active

    assert sheet.max_row == 3  # header + two documents


def test_browsing_metadata_needs_the_application_view_permission(client):
    client.headers.update({"REMOTE-USER": "nobody"})

    assert client.get("/file-metadata").status_code == 403
    assert client.get("/file-metadata/export").status_code == 403
