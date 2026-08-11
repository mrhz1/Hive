"""Where de-identification records what it produced.

Two separate places, deliberately: `file_metadata` holds what a document
arrived carrying, and the facts this system generates go into the output
file's own metadata. See app/embed.py.
"""
import pathlib

import pytest

from app import deid
from app.embed import embed_metadata, generated_facts
from conftest import minimal_patient

GENERATED_KEYS = (
    "deidentified",
    "deidentified_at",
    "deidentified_by",
    "deidentified_file_name",
    "deidentified_file_type",
    "deidentified_method",
    "patient_id",
)


def _file_with_patient(client, name="scan.pdf", data=b"%PDF-1.4 fake"):
    patient_id = client.post("/patients", json=minimal_patient()).json()["id"]
    application_id = client.post(
        "/applications", json={"patient_id": patient_id}
    ).json()["id"]
    record = client.post(
        f"/applications/{application_id}/files",
        files=[("files", (name, data, "application/pdf"))],
    ).json()[0]
    return patient_id, record


def _row(record):
    class Row:
        id = record["id"]
        application_id = record["application_id"]

    return Row()


# ------------------------------------------- what does NOT go in the row


def test_generated_facts_stay_out_of_the_metadata_row(as_admin, storage_root):
    _, record = _file_with_patient(as_admin)

    deid._record_deid_metadata(_row(record), pathlib.Path("/out/x_deid.pdf"))

    metadata = as_admin.get(f"/files/{record['id']}/metadata").json()["metadata"]
    for key in GENERATED_KEYS:
        assert key not in metadata, key


def test_the_original_extraction_is_left_untouched(as_admin, storage_root, cursor):
    """Recording the output must not disturb what was read off the input."""
    _, record = _file_with_patient(as_admin)

    for row in cursor.store["file_metadata"]:
        if row["file_id"] == record["id"]:
            row["metadata"] = '{"Author": "Dr Grant"}'

    deid._record_deid_metadata(_row(record), pathlib.Path("/out/x_deid.pdf"))

    metadata = as_admin.get(f"/files/{record['id']}/metadata").json()["metadata"]
    assert metadata == {"Author": "Dr Grant"}


def test_a_missing_output_file_does_not_raise(as_admin, storage_root):
    """The redaction already succeeded; a failed annotation must not turn that into a failed run."""
    _, record = _file_with_patient(as_admin)

    # No such path -- embed_metadata logs and returns, it does not raise.
    deid._record_deid_metadata(_row(record), pathlib.Path("/out/nope_deid.pdf"))


def test_a_patient_lookup_failure_does_not_raise(
    as_admin, storage_root, monkeypatch
):
    def boom(*args, **kwargs):
        raise RuntimeError("hive is down")

    monkeypatch.setattr(deid, "_patient_id_for", boom)
    _, record = _file_with_patient(as_admin)

    deid._record_deid_metadata(_row(record), pathlib.Path("/out/x_deid.pdf"))


# ------------------------------------------------ what goes into the file


def test_the_facts_are_added_to_a_pdf_without_replacing_it(tmp_path):
    """The pipeline has already de-identified these fields in place, so
    'Author: <PERSON>' is the de-identified value, not a leak. Overwriting
    it would throw away the fact that a person was named there."""
    fitz = pytest.importorskip("fitz")

    path = tmp_path / "out_deid.pdf"
    document = fitz.open()
    document.new_page()
    document.set_metadata(
        {"author": "<PERSON>", "title": "Discharge summary", "producer": "Acme 4.1"}
    )
    document.save(str(path))
    document.close()

    written = embed_metadata(
        path, "pdf", generated_facts(patient_id="A7K2P9", output_name=path.name)
    )
    assert written == "pdf"

    reopened = fitz.open(str(path))
    info = reopened.metadata
    reopened.close()

    # What the pipeline left is still there.
    assert info["author"] == "<PERSON>"
    assert info["title"] == "Discharge summary"
    assert info["producer"] == "Acme 4.1"
    # And our own facts are alongside it.
    assert "patient_id=A7K2P9" in info["keywords"]
    assert "deidentified=yes" in info["keywords"]


def test_the_facts_are_written_into_a_dicom(tmp_path):
    pydicom = pytest.importorskip("pydicom")
    from pydicom.dataset import Dataset, FileMetaDataset
    from pydicom.uid import ExplicitVRLittleEndian

    path = tmp_path / "study_deid.dcm"
    dataset = Dataset()
    dataset.PatientName = "Doe^Jane"
    dataset.file_meta = FileMetaDataset()
    dataset.file_meta.TransferSyntaxUID = ExplicitVRLittleEndian
    dataset.file_meta.MediaStorageSOPClassUID = "1.2.840.10008.5.1.4.1.1.7"
    dataset.file_meta.MediaStorageSOPInstanceUID = "1.2.3.4"
    dataset.save_as(str(path), enforce_file_format=True)

    written = embed_metadata(
        path, "dcm", generated_facts(patient_id="A7K2P9", output_name=path.name)
    )
    assert written == "dicom"

    reread = pydicom.dcmread(str(path), force=True)
    assert reread.PatientIdentityRemoved == "YES"
    assert "patient_id=A7K2P9" in list(reread.DeidentificationMethod)


def test_a_dicom_keeps_the_method_the_pipeline_already_set(tmp_path):
    pydicom = pytest.importorskip("pydicom")
    from pydicom.dataset import Dataset, FileMetaDataset
    from pydicom.uid import ExplicitVRLittleEndian

    path = tmp_path / "study_deid.dcm"
    dataset = Dataset()
    dataset.DeidentificationMethod = "Hive OCR/NER pixel and tag de-identification"
    dataset.file_meta = FileMetaDataset()
    dataset.file_meta.TransferSyntaxUID = ExplicitVRLittleEndian
    dataset.file_meta.MediaStorageSOPClassUID = "1.2.840.10008.5.1.4.1.1.7"
    dataset.file_meta.MediaStorageSOPInstanceUID = "1.2.3.4"
    dataset.save_as(str(path), enforce_file_format=True)

    embed_metadata(path, "dcm", generated_facts(patient_id="A7K2P9"))

    methods = list(pydicom.dcmread(str(path), force=True).DeidentificationMethod)
    assert "Hive OCR/NER pixel and tag de-identification" in methods
    assert "patient_id=A7K2P9" in methods


def test_the_facts_are_added_to_a_word_document_without_replacing_it(tmp_path):
    docx = pytest.importorskip("docx")

    path = tmp_path / "letter_deid.docx"
    document = docx.Document()
    document.add_paragraph("body")
    document.core_properties.author = "<PERSON>"
    document.core_properties.keywords = "cardiology"
    document.save(str(path))

    written = embed_metadata(
        path, "docx", generated_facts(patient_id="A7K2P9", output_name=path.name)
    )
    assert written == "word"

    properties = docx.Document(str(path)).core_properties
    assert properties.author == "<PERSON>"
    assert properties.keywords == "cardiology"
    assert "patient_id=A7K2P9" in properties.comments


def test_an_unsupported_format_is_skipped_quietly(tmp_path):
    path = tmp_path / "notes.txt"
    path.write_text("hello")

    assert embed_metadata(path, "txt", generated_facts(patient_id="A7K2P9")) is None
    assert path.read_text() == "hello"


def test_a_corrupt_file_is_logged_not_raised(tmp_path):
    path = tmp_path / "broken_deid.pdf"
    path.write_bytes(b"not really a pdf")

    assert embed_metadata(path, "pdf", generated_facts(patient_id="A7K2P9")) is None


# ------------------------------------------------------------ xlsx export


def _seed_metadata(cursor, record, values):
    """Write straight to the store: the API no longer merges into a row."""
    import json

    for row in cursor.store["file_metadata"]:
        if row["file_id"] == record["id"]:
            row["metadata"] = json.dumps(values)


def _read_workbook(content):
    import io

    from openpyxl import load_workbook

    sheet = load_workbook(io.BytesIO(content)).active
    return [[cell.value for cell in row] for row in sheet.iter_rows()]


def test_export_is_a_real_workbook(as_admin, storage_root, cursor):
    _, record = _file_with_patient(as_admin)
    _seed_metadata(cursor, record, {"Author": "Dr Grant", "Pages": "3"})

    response = as_admin.get(f"/files/{record['id']}/metadata/export")

    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["content-type"]
    assert ".xlsx" in response.headers["content-disposition"]
    # A real xlsx is a zip; a CSV renamed would not be.
    assert response.content[:2] == b"PK"

    rows = _read_workbook(response.content)
    assert rows[0] == ["Field", "Value"]
    assert ["Author", "Dr Grant"] in rows
    assert ["Pages", "3"] in rows


def test_export_honours_the_filter(as_admin, storage_root, cursor):
    """Exporting everything when the screen shows two rows would not be the data the user filtered down to."""
    _, record = _file_with_patient(as_admin)
    _seed_metadata(
        cursor, record, {"Author": "Dr Grant", "Pages": "3", "Producer": "Acme"}
    )

    response = as_admin.get(
        f"/files/{record['id']}/metadata/export", params={"fields": "Author,Pages"}
    )

    rows = _read_workbook(response.content)
    fields = [row[0] for row in rows[1:]]
    assert sorted(fields) == ["Author", "Pages"]
    assert "Producer" not in fields


def test_export_keeps_long_numbers_as_text(as_admin, storage_root, cursor):
    """Excel turns a 16-digit serial into 1.78633E+15 if it is left to guess, and the value does not survive a round trip."""
    _, record = _file_with_patient(as_admin)
    _seed_metadata(cursor, record, {"serial": "1786329822402000"})

    response = as_admin.get(f"/files/{record['id']}/metadata/export")

    rows = _read_workbook(response.content)
    value = next(row[1] for row in rows[1:] if row[0] == "serial")
    assert value == "1786329822402000"
    assert isinstance(value, str)


def test_export_for_a_file_with_no_metadata_row_is_a_404(as_admin, storage_root):
    response = as_admin.get("/files/nope/metadata/export")
    assert response.status_code == 404
